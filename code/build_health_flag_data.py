#!/usr/bin/env python
"""Assemble the data side of the health-expenditure flag decomposition.

Reads three blocks from data/DATA_GR.xlsx and writes one tidy CSV plus a
companion npz with the population age-share matrix:

  * "Input check" — Eurostat hlth_sha11_hf, health care expenditure by financing
    scheme, Greece, million euro.  Columns: All financing schemes (current health
    expenditure, CHE), Government/compulsory schemes, ..., Household out-of-pocket.
    Coverage ratio kappa_data = Government / All.  Available 2009-2023.
  * "Input" — nominal GDP, Greece, bn euro (Eurostat).  Denominator for the /GDP
    ratios.
  * "Population by age" — Eurostat population on 1 January, single year of age,
    Greece.  Used for data population age shares over the model age range
    (real ages 25-84 = model ages 0..T-1).

The data age-cost index Abar_data_t = sum_j s^d_{j,t} a(j) uses the SAME age
profile a(j) the model uses (m_age_profile in the config), so the demographics
factor isolates age-composition differences, not a different profile.

Output columns (data/health_flag_GR.csv), one row per year with health data:
    year
    gov_health_meur, che_meur, oop_meur   raw million-euro levels
    gdp_meur                              nominal GDP, million euro
    kappa_data                            Government / All  (coverage)
    gov_health_gdp                        Gov / GDP         (the target g^data)
    che_gdp                               CHE / GDP         (total health share)
    oop_gdp                               OOP / GDP
    Abar_data                             data age-cost index (model a(j))
    pop_2584                              population aged 25-84 (persons)

Companion data/health_flag_age_shares_GR.npz:
    year (n,), model_age (T,), real_age (T,), share (n, T)  data age shares s^d_{j,t}
"""
import argparse
import json
import numpy as np
import openpyxl


def _as_year(v):
    """Parse a cell into a calendar year, tolerating text-stored years."""
    try:
        y = int(float(v))
    except (TypeError, ValueError):
        return None
    return y if 1900 <= y <= 2100 else None


def _find_year_rows(ws, year_col, first_data_row):
    out = {}
    for row in ws.iter_rows(min_row=first_data_row, values_only=True):
        y = _as_year(row[year_col])
        if y is not None:
            out[y] = row
    return out


def read_health(wb):
    """Input check: year col1, All col2, Gov col3, OOP col5 (million euro)."""
    ws = wb["Input check"]
    rows = _find_year_rows(ws, year_col=1, first_data_row=9)
    health = {}
    for y, row in rows.items():
        allsch, gov, oop = row[2], row[3], row[5]
        if isinstance(allsch, (int, float)) and isinstance(gov, (int, float)):
            health[y] = dict(che=float(allsch), gov=float(gov),
                             oop=float(oop) if isinstance(oop, (int, float)) else np.nan)
    return health


def read_gdp(wb):
    """Input: year col2, nominal GDP col12 (bn euro) -> million euro."""
    ws = wb["Input"]
    rows = _find_year_rows(ws, year_col=2, first_data_row=9)
    gdp = {}
    for y, row in rows.items():
        g = row[12]
        if isinstance(g, (int, float)):
            gdp[y] = float(g) * 1000.0   # bn -> million
    return gdp


def read_population(wb, n_ages, real_age0=25):
    """Population by age: year col1, age a in col (3+a). Model age j -> real age
    real_age0+j -> column 3 + real_age0 + j.  Returns {year: np.array(n_ages)}."""
    ws = wb["Population by age"]
    rows = list(ws.iter_rows(values_only=True))
    hdr_i = next(i for i, r in enumerate(rows)
                 if r and any(isinstance(c, str) and "Less than 1" in c for c in r))
    col0 = 3 + real_age0                       # column of the first model age
    pop = {}
    for r in rows[hdr_i + 1:]:
        y = _as_year(r[1])
        if y is None:
            continue
        vals = r[col0:col0 + n_ages]
        if all(isinstance(v, (int, float)) for v in vals):
            pop[y] = np.array([float(v) for v in vals])
    return pop


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--xlsx", default="../data/DATA_GR.xlsx")
    ap.add_argument("--config", default="calibration_input_GR.json")
    ap.add_argument("--out-csv", default="data/health_flag_GR.csv")
    ap.add_argument("--out-npz", default="data/health_flag_age_shares_GR.npz")
    args = ap.parse_args()

    cfg = json.load(open(args.config))
    a = np.asarray(cfg["m_age_profile"], dtype=float)
    T = len(a)

    wb = openpyxl.load_workbook(args.xlsx, read_only=True, data_only=True)
    health = read_health(wb)
    gdp = read_gdp(wb)
    pop = read_population(wb, n_ages=T, real_age0=25)

    years = sorted(y for y in health if y in gdp and y in pop)
    if not years:
        raise SystemExit("no overlapping years across health / GDP / population")

    rows_out = []
    share_mat = np.zeros((len(years), T))
    for i, y in enumerate(years):
        h = health[y]
        g = gdp[y]
        p = pop[y]
        s = p / p.sum()
        share_mat[i] = s
        Abar = float(np.sum(s * a))
        rows_out.append(dict(
            year=y,
            gov_health_meur=h["gov"], che_meur=h["che"], oop_meur=h["oop"],
            gdp_meur=g,
            kappa_data=h["gov"] / h["che"],
            gov_health_gdp=h["gov"] / g,
            che_gdp=h["che"] / g,
            oop_gdp=(h["oop"] / g) if np.isfinite(h["oop"]) else np.nan,
            Abar_data=Abar,
            pop_2584=float(p.sum()),
        ))

    cols = ["year", "gov_health_meur", "che_meur", "oop_meur", "gdp_meur",
            "kappa_data", "gov_health_gdp", "che_gdp", "oop_gdp",
            "Abar_data", "pop_2584"]
    with open(args.out_csv, "w") as f:
        f.write(",".join(cols) + "\n")
        for r in rows_out:
            f.write(",".join(
                f"{r[c]:.8g}" if isinstance(r[c], float) else str(r[c]) for c in cols
            ) + "\n")

    np.savez(args.out_npz,
             year=np.array(years), model_age=np.arange(T),
             real_age=25 + np.arange(T), share=share_mat)

    print(f"Saved {args.out_csv}  ({len(years)} years: {years[0]}-{years[-1]})")
    print(f"Saved {args.out_npz}")
    print(f"  {'year':>5} {'kappa':>7} {'gov/Y':>7} {'che/Y':>7} {'Abar':>7}")
    for r in rows_out:
        print(f"  {r['year']:>5} {r['kappa_data']:>7.3f} {r['gov_health_gdp']:>7.4f} "
              f"{r['che_gdp']:>7.4f} {r['Abar_data']:>7.4f}")


if __name__ == "__main__":
    main()
